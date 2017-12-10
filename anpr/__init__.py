import datetime
import glob
import os
import argparse
import re
import functools

import openpyxl
import psycopg2 as psy
from psycopg2 import sql
from fastkml import kml

from anpr import filters
from anpr import groups
from anpr import stats


UNINTERESTING_SHEETS = (
    "Front Cover", "Contents Page", "Location Plan", "Summary", "Trip Arrays")
CAMERA_ID_CELL = "C4"

CAMERA_START=1
CAMERA_END=97
DATA_START_ROW = 12
DATA_START_COL = 2
DATA_END_COL = 6

CHAIN_DIRECTION_REGEX = re.compile("^.*?_(N|E|S|W|IN|OUT)>")
DESTINATIONS_REGEX = re.compile(r">(.*?)_(N|E|S|W|IN|OUT)\(([\d\.]+)\)")


def make_journeys_table(dbname, password):
    conn = psy.connect("dbname={} password={}".format(dbname, password))
    cur = conn.cursor()
    cur.execute("CREATE TABLE journeys (journey_id SERIAL PRIMARY KEY"+
    ", timestamp timestamp"+
    ", class text"+
    ", total_trip_time interval"+
    ", chain text"+
    ", trip_destinations_and_time text"+
    ", journey_end_time timestamp);"
    )
    conn.commit()


class DataLoader:
    def __init__(self, spreadsheet_path, db_connection):
        wb = openpyxl.load_workbook(
            filename=spreadsheet_path, read_only=True)
        self.wb = wb
        self.conn = db_connection

    def load(self):
        camera_sheets = [sheet for sheet in self.wb.worksheets
                         if sheet.title not in UNINTERESTING_SHEETS]
        for sheet in camera_sheets:
            # Sanity check -- can be pretty sure we're loading camera data.
            camera_name = sheet.title
            check = str(sheet[CAMERA_ID_CELL].value)
            if camera_name not in (check, "0"+check):
                raise ValueError(
                    "Sheet titled {!r} doesn't look like camera data".format(
                        camera_name))

            print("Loading trips starting at camera {}".format(sheet.title))
            # In case sheets report an incorrect size.
            sheet.max_row = None
            sheet.max_column = None
            cursor = self.conn.cursor()
            for row in sheet.iter_rows(
                    min_row=DATA_START_ROW,
                    min_col=DATA_START_COL, max_col=DATA_END_COL):
                self.load_chain(row, camera_name, cursor)

        self.conn.commit()

    def load_chain(self, row, camera_name, cursor):
        timestamp, veh_class, _tot_mins, chain, details = row

        if timestamp.value is None:
            print(
                "Empty row in '{}' cell {}:{}".format(
                camera_name, timestamp.column, timestamp.row))
            return

        if not isinstance(timestamp.value, datetime.datetime):
            raise ValueError(
                "Expected a datetime from cell {}:{}, was {}".format(
                    timestamp.column, timestamp.row, type(timestamp.value)))
        timestamp = timestamp.value
        veh_class = veh_class.value

        cursor.execute(
            "INSERT INTO vehicles (class) VALUES (%s)"
            "RETURNING id;",
            (veh_class,)
        )
        vehicle_id = cursor.fetchone()

        initial_camera = camera_name
        match = CHAIN_DIRECTION_REGEX.match(chain.value)
        if not match:
            raise ValueError(
                "Could not extract the initial direction from the chain in "
                "cell {}:{}: {!r}".format(chain.column, chain.row, chain.value))
        initial_direction = match.group(1)
        cursor.execute(
            "INSERT INTO captures (camera, vehicle, direction, ts)"
            "VALUES (%s, %s, %s, %s);",
            (initial_camera, vehicle_id, initial_direction, timestamp)
        )

        found_one = False
        next_ts = timestamp
        for match in DESTINATIONS_REGEX.finditer(details.value):
            found_one = True
            next_camera = match.group(1)
            next_direction = match.group(2)
            next_duration = float(match.group(3))
            next_ts = next_ts + datetime.timedelta(minutes=next_duration)
            cursor.execute(
                "INSERT INTO captures (camera, vehicle, direction, ts)"
                "VALUES (%s, %s, %s, %s);",
                (next_camera, vehicle_id, next_direction, next_ts)
            )
        assert found_one, (
            "No trip details found in cell {}:{}".format(
            details.column, details.row))

    def load_journey(self, row):
        '''
        load the given journey entry into the database
        '''
        if row[0].value is not None:
            journey_id = self.add_journey_entry(row)
            for site in set(row[3].value.split(">")):
                if not self.table_exists("s"+site):
                    self.create_site_set_table("s"+site)
                self.add_to_site_set("s"+site, journey_id)

    def add_journey_entry(self, row):
        cur = self.conn.cursor()
        trip_time = datetime.timedelta(minutes=row[2].value)
        if type(row[0].value) == str:
            end_time = datetime.datetime.strptime(row[0].value,"%d/%m/%Y %H:%M:%S") + trip_time
        else:
            end_time = row[0].value + trip_time
        cur.execute("INSERT INTO journeys (timestamp, class, total_trip_time, chain, trip_destinations_and_time, journey_end_time)" +
        "VALUES (%s, %s, %s, %s, %s, %s) RETURNING journey_id;", [row[0].value,
        row[1].value, trip_time, row[3].value, row[4].value, end_time])
        journey_id = cur.fetchone()

        self.conn.commit()
        return journey_id

    def table_exists(self, table_name):
        cur = self.conn.cursor()
        cur.execute("select relname from pg_class where relname = %s;", [table_name])
        if cur.fetchone():
            return True
        else:
            return False

    def create_site_set_table(self, site):
        cur = self.conn.cursor()
        s = sql.SQL("CREATE TABLE {} (journey_id serial UNIQUE REFERENCES"+
        " journeys);").format(
            sql.Identifier(site))
        cur.execute(s)
        self.conn.commit()

    def add_to_site_set(self, site, journey_id):
        cur = self.conn.cursor()
        s = sql.SQL("INSERT INTO {} (journey_id) VALUES (%s);").format(sql.Identifier(site))
        cur.execute(s, [journey_id])

        self.conn.commit()

CHAIN_COLUMN_INDEX = 4
CHAIN_TIME_COLUMN_INDEX = 5
CLASS_COLUMN_INDEX = 2
TIMESTAMP_COLUMN_INDEX = 1
TOTAL_TIME_COLUMN_INDEX = 3

def compose(functions):
    return functools.reduce(lambda f, g: lambda x: f(g(x)), functions, lambda x: x)

class DataSearcher(object):
    def __init__(self, dbname, db_password, filter_lst=[], group_lst=[], stats_lst=[]):
        self.conn = psy.connect("dbname={} password={}".format(dbname, db_password))
        for fil in filter_lst:
            assert(isinstance(fil, filters.FilterBase))
        self.filters = filter_lst
        #compose all the fine pass filters into one function
        #TODO check that the order is preserved
        self.fine_pass = compose([fil.fine_pass for fil in filter_lst])

        for group in group_lst:
            assert(isinstance(group, groups.GroupBase))
        self.group = compose([group.group for group in group_lst])

        for stat in stats_lst:
            assert(isinstance(stat, stats.BaseStats))
        self.stats = stats_lst

    def get_and_filter(self):
        '''
        Go to the DB and apply the filters
        '''
        cur = self.conn.cursor()
        sql_filters = sql.SQL(" AND ").join([fil.coarse_pass() for fil in self.filters])
        if self.filters:
            cur.execute(sql.SQL("SELECT * from journeys where {};").format(sql_filters))
        else:
            #no filter means no WHERE
            cur.execute("SELECT * from journeys;")
        return self.fine_pass(cur)

    def combined(self):
        '''
        get the results from the db, apply the filters,
        group then get the statistics for each group of rows
        '''
        groups = self.group(list(self.get_and_filter()))
        return self.apply_stats(groups)

    def stat_headers(self):
        out = []
        for stat in self.stats:
            out += stat.stat_descriptions()
        return out

    def apply_stats(self, group_or_rows):
        if isinstance(group_or_rows, list):
            stat_lists = [stats.make_stats(group_or_rows) for stats in self.stats]
            return [stat for sublist in stat_lists for stat in sublist]
        elif isinstance(group_or_rows, dict):
            return {key: self.apply_stats(value) for key, value in group_or_rows.items()}
        else:
            raise Exception("Unknown group type:{}".format(type(group_or_rows)))


def main():
    args = parse_args()
    if args.command_name == "load":
        do_load_command(args)
    elif args.command_name == "create":
        do_create_command(args)

def parse_args():
    parser = argparse.ArgumentParser(
        description="Script for loading data from anpr spreadsheets into a db")
    parser.add_argument(
        "--dbname", required=True, help="name of the db to create")
    parser.add_argument(
        "--password", required=True, help="password to the database")
    parser.add_argument(
        "--user", required=True,
        help="the username used to access the database")
    subparsers = parser.add_subparsers(dest="command_name")

    load = subparsers.add_parser(
        "load", help="Load data into an existing database")
    load.add_argument(
        "xlsx_dir", help="path to the directory where the spreadsheets are")

    create = subparsers.add_parser(
        "create", help="Create the database")
    create.add_argument(
        "cameras", help="Path to the file that defines the cameras")

    args = parser.parse_args()
    return args

def do_load_command(args):
    for spreadsheet_path in glob.glob(
            os.path.join(os.path.abspath(args.xlsx_dir), "*.xlsx")):
        print("loading {!r}...".format(spreadsheet_path))
        DataLoader(
            spreadsheet_path, db_connection=make_connection(args)).load()
        print("loaded")

def do_create_command(args):
    """Create the initial database tables.

    Pre-requisites: the database must already exist, and the PostGIS extension
    must already be installed. Doing this requires admin-ish rights, so this
    function won't try to do it automatically.

    The location of the ANPR cameras is not contained directly within the
    spreadsheets alongside the trip data. Rather, that data is exposed as a
    set of location markers in Google Maps (linked to in the Location Plan
    within each spreadsheet).
    We assume that this location data has been exported from Google Maps and
    saved as a KML file.  This function parses the KML file to get a location
    and description for each camera.

    We create and populate the camera location table using this information.
    We also create other empty tables that will hold the trip data (loaded in
    a later stage).
    """

    with open(args.cameras, 'rb') as infile:
        s=infile.read()
    k = kml.KML()
    k.from_string(s)

    doc = list(k.features())[0]
    placemarks = doc.features()

    cameras = []
    for place in placemarks:
        name = place.name
        # 1-9 -> 01-09
        try:
            num = int(name)
        except Exception:
            pass
        else:
            if num < 10:
                name = "0"+name

        if place.geometry.geom_type != "Point":
            raise ValueError(
                "Placemark {!r} has a geometry type of {!r} "
                "(expected Point)".format(name, place.geometry.geom_type))
        # Weird unpacking -- it's a tuple of a tuple.
        (x, y, z), = place.geometry.coords
        description = [e.value for e in place.extended_data.elements
                       if e.name == "Description"][0]
        cameras.append((name, description, x, y))

    with make_connection(args) as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT postgis_full_version();")
            ver, = cur.fetchone()
            if not "POSTGIS=" in ver:
                raise RuntimeError("DB does not seem to have PostGIS installed")

        with conn.cursor() as cur:
            # (Re)make the table of cameras.
            cur.execute(
                "DROP TABLE IF EXISTS cameras CASCADE;"
            )
            cur.execute(
                "CREATE TABLE cameras ("
                "id varchar(10) PRIMARY KEY, "
                "description text, "
                "location geometry NOT NULL"
                ");"
            )
            # And populate it.
            for name, description, x, y in cameras:
                ewkt = "SRID=4326;POINT({:.6f} {:.6f})".format(x, y)
                cur.execute(
                    "INSERT INTO cameras (id, description, location) "
                    "VALUES (%s, %s, ST_GeomFromEWKT(%s));",
                    (name, description, ewkt)
                )

        with conn.cursor() as cur:
            # (Re)make the vehicle class enum type.
            cur.execute(
                "DROP TYPE IF EXISTS veh_class CASCADE;"
            )
            cur.execute(
                "CREATE TYPE veh_class AS ENUM ("
                    "'Bus_Coach', "
                    "'Car', "
                    "'LGV<3.5T', "
                    "'Motorcycle', "
                    "'OGV1', "
                    "'OGV2', "
                    "'Other', "
                    "'Taxi'"
                ");"
            )
            # (Re)make the vehicle table.
            cur.execute(
                "DROP TABLE IF EXISTS vehicles CASCADE;"
            )
            cur.execute(
                "CREATE TABLE vehicles ("
                "id integer PRIMARY KEY GENERATED ALWAYS AS IDENTITY, "
                "class veh_class NOT NULL"
                ");"
            )

        with conn.cursor() as cur:
            # (Re)make the direction enum type.
            cur.execute(
                "DROP TYPE IF EXISTS direction CASCADE;"
            )
            cur.execute(
                "CREATE TYPE direction AS ENUM ("
                    "'N', "
                    "'S', "
                    "'E', "
                    "'W', "
                    "'IN', "
                    "'OUT'"
                ");"
            )
            # (Re)make the capture table.
            cur.execute(
                "DROP TABLE IF EXISTS captures CASCADE;"
            )
            cur.execute(
                "CREATE TABLE captures ("
                "id integer PRIMARY KEY GENERATED ALWAYS AS IDENTITY, "
                "camera varchar(10) NOT NULL REFERENCES cameras (id), "
                "vehicle integer NOT NULL REFERENCES vehicles (id), "
                "direction direction NOT NULL, "
                "ts timestamptz NOT NULL"
                ");"
            )

def make_connection(args):
    conn = psy.connect(
        dbname=args.dbname, user=args.user, password=args.password)
    return conn


if __name__=="__main__":
    main()
